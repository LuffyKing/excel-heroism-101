from openpyxl.styles import PatternFill
import os, openpyxl, re
from fuzzywuzzy import fuzz


def generateSetOfActiveVendors(sheet):
    """
    Input: an excel sheet, an empty set
    Output: a set of active vendors
    Description:
    This function iterates over the cells in the vendor column , and then adds
    the vendor codes to the input set.
    """
    activeVendorsSets = set()
    startColumn, startRow = findVendorCodeColumnAndRow(sheet)
    for arow in range(startRow, sheet.max_row+1):
        if sheet[startColumn+str(arow)].value:
            if re.search(pattern=r'^VDR\d+$', string=str(sheet[startColumn+str(arow)].value)):
                activeVendorsSets.add(sheet[startColumn+str(arow)].value)
    return activeVendorsSets


def findVendorCodeColumnAndRow(sheet):
    """
    Input: an excel sheet
    Output: a cell letter and the row number
    Description:
    This function iterates over cells in the provided excel sheet in a row by row
    manner and gets the first vendor cell, which will be used as the starting point.
    """

    for aTupleOfCells in sheet.rows:
        for acell in aTupleOfCells:
            if acell.value:
                if re.search(pattern=r'^VDR\d+$', string=str(acell.value)):  # VDR87686652
                    return acell.column, acell.row  # 'A, 2'

def findStatusColumn(sheet):
    """
    Input: an excel sheet
    Output: a cell letter
    Description:
    This function iterates over cells in the provided excel sheet in a row by row
    manner and gets the status column, which will be used as the starting point.
    """
    for aTupleOfCells in sheet.rows:
        for acell in aTupleOfCells:
            if acell.value:
                if acell.value.upper() == 'STATUS':
                    return acell.column  # 'A'
                elif fuzz.partial_token_set_ratio('STATUS', str(acell.value)) > 80: #Tokenise and compare based on intersecting substrings and remainders
                    return acell.column  # 'A'


def findActiveVendors(activeVendorsSet, sheet):
    """
    Input: an excel sheet, a set of active vendors
    Output: None
    Description:
    This function iterates over the cells in the vendor column , and then
    highlights the row of the matching cell in the vendor column.
    """
    redFill = PatternFill(start_color='ff4c4c', end_color='ff4c4c', fill_type='solid')
    startColumn, startRow = findVendorCodeColumnAndRow(sheet)
    activeColumn = findStatusColumn(sheet)
    for rowIndex in range(startRow, sheet.max_row+1):
        if sheet[startColumn+str(rowIndex)].value in activeVendorsSet:
            sheet[activeColumn+str(rowIndex)].value = 'Active'
            for columnIndex in range(1, sheet.max_column+1):
                sheet.cell(row=rowIndex, column=columnIndex).fill = redFill
        else:
            sheet[activeColumn+str(rowIndex)].value = 'Inactive'





def runTask(activeVendorsSet):
    """
    Input:a set of active vendors
    Output: None
    Description:
    This function iterates over vendor excel files in the folder specified below
    ,finds the active vendors in each excel file and also highlights them.
    It does this by calling the findActiveVendors function for each sheet.
    """
    for afile in os.listdir('/Users/Damola/Documents/Pycon2017/Unique Identifier/vendor'):
        if '$' not in afile and '.xlsx' in afile:
            wb = openpyxl.load_workbook(os.path.join('/Users/Damola/Documents/Pycon2017/Unique Identifier/vendor', afile))
            for asheet in wb:
                findActiveVendors(activeVendorsSet, asheet)
            wb.save('/Users/Damola/Documents/Pycon2017/Unique Identifier/result/altered %s' % afile)
