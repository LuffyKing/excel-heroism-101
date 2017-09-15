from fuzzywuzzy import process
import openpyxl
from openpyxl.styles import PatternFill, Border,Side


def createDicts():
    """
    Input: none
    Output: comparison promoters dictionary and master promoters dictionary
    Description:
    This function iterates over the sheets in  the master and comp excel files and puts
    their values in a dictionaries by state.
    """
    masterListWb = openpyxl.load_workbook('/Users/Damola/'
                                          'Documents/Pycon2017/Non-Unique Identifier/MASTER SHEET 2017.xlsx')
    compListWb = openpyxl.load_workbook('/Users/Damola/Documents/Pycon2017/Non-Unique Identifier/SHEET 2.xlsx')
    masterPromotersDict = {}
    for stateSheet in masterListWb:
        masterPromotersDict[stateSheet.title] = {}

        for arow in range(1, stateSheet.max_row+1):
            if stateSheet['B'+str(arow)].value and 'PROMOTER' not in stateSheet['B'+str(arow)].value.upper():
                masterPromotersDict[stateSheet.title][stateSheet['B'+str(arow)].value] = stateSheet['C'+str(arow)].value
    compPromotersDict = {}
    for stateSheet in compListWb:
        compPromotersDict[stateSheet.title.upper()] = {}

        for arow in range(1, stateSheet.max_row + 1):
            if stateSheet['A'+str(arow)].value and 'VENDOR' not in stateSheet['A'+str(arow)].value.upper():
                compPromotersDict[stateSheet.title.upper()][stateSheet['A'+str(arow)].value] = \
                                                                                        stateSheet['B'+str(arow)].value
    return compPromotersDict, masterPromotersDict


def consolidateWb(compPromotersDict, masterPromotersDict):
    """
    Input: comparison promoters dictionary and master promoters dictionary
    Output: none
    Description:
    This function iterates over the comparison promoters dictionary and master promoters dictionary, and compares
    promoter names across the two dictionaries using fuzzywuzzy
    """
    endresult = []
    for state in compPromotersDict:
        for promoter in compPromotersDict[state]:
            choices = masterPromotersDict[state.upper()].keys()#promoters in master file
            
            answer = process.extractOne(promoter, choices)
            if answer[1] >= 80:
                endresult.append([promoter,
                                  compPromotersDict[state][promoter],
                                  answer[0],
                                  masterPromotersDict[state.upper()][answer[0]],
                                  state.upper()])

    endwb = openpyxl.Workbook()
    sheet = endwb.active
    headerRow = ['NAME MASTERLIST',
                 'SALARY MASTERLIST',
                 'NAME',
                 'SALARY',
                 'STATE']
    lightBlueFill = PatternFill(start_color='add8e6', end_color='add8e6', fill_type='solid')
    lightYellowFill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    thick = Side(border_style="thick")
    thin = Side(border_style="thin")
    thickBorder = Border(top=thick, left=thick, right=thick, bottom=thick)
    thinBorder = Border(thin, thin, thin, thin)

    for num, aHeader in enumerate(headerRow, start=1):
        sheet.cell(row=1, column=num).value = aHeader
        sheet.cell(row=1, column=num).fill = lightBlueFill
        sheet.cell(row=1, column=num).border = thickBorder

    for num, aresult in enumerate(endresult, start=2):
        for index, element in enumerate(aresult):
            sheet.cell(row=num, column=index+1).value = element
            sheet.cell(row=num, column=index + 1).fill = lightYellowFill
            sheet.cell(row=num, column=index + 1).border = thinBorder

    endwb.save('/Users/Damola/Documents/Pycon2017/Non-Unique Identifier/Consolidated.xlsx')
